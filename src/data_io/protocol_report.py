"""Экспорт отчёта по протоколам кошельков в Excel."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.db.dao import Dao
from src import logger

_thin = Side(style="thin", color="B0B0B0")
_sep = Side(style="medium", color="2F5496")   # толстая линия-разделитель между кошельками
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_HEAD_FILL = PatternFill("solid", start_color="2F5496")
_HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_BODY = Font(name="Arial", size=10)
# чередующиеся заливки групп-кошельков (для визуального разделения)
_BANDS = (PatternFill("solid", start_color="FFFFFF"), PatternFill("solid", start_color="E9F0FB"))


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def export_protocol_report(dao: Dao, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Лист 1: использование протоколов по кошелькам
    ws = wb.active
    ws.title = "protocols"
    headers = ["Кошелёк", "Метка", "AGW-адрес", "Протокол", "Chain", "Тип позиции", "USD", "Проверено"]
    ws.append(headers)
    _style_header(ws, len(headers))
    rows = dao.protocol_report_rows()
    for r in rows:
        ws.append([
            r["address"], r["label"] or "", r["agw_address"] or "", r["protocol"], r["chain"],
            r["item_types"] or "", _num(r["net_usd"]),
            _short_ts(r["checked_at"]),
        ])
    # Разделение по кошелькам: чередующаяся заливка на группу + толстая линия сверху при смене кошелька.
    prev_addr = None
    group_idx = -1
    for i, (row, r) in enumerate(zip(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)), rows)):
        addr = r["address"]
        new_group = addr != prev_addr
        if new_group:
            group_idx += 1
            prev_addr = addr
        fill = _BANDS[group_idx % 2]
        for cell in row:
            cell.font = _BODY
            cell.fill = fill
            # первая строка нового кошелька (кроме самой первой) — толстая верхняя граница-разделитель
            top = _sep if (new_group and i > 0) else _thin
            cell.border = Border(left=_thin, right=_thin, top=top, bottom=_thin)
            cell.alignment = Alignment(vertical="center")
        row[6].number_format = "$#,##0.00"
    widths = [46, 16, 46, 22, 8, 24, 12, 20]
    for col, w in zip("ABCDEFGH", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # Лист 2: сводка по кошелькам (сколько протоколов, сумма)
    ws2 = wb.create_sheet("summary")
    ws2.append(["Кошелёк", "Метка", "AGW-адрес", "Протоколов", "USD в протоколах"])
    _style_header(ws2, 5)
    agg: dict[tuple, list] = {}
    for r in rows:
        key = (r["address"], r["label"] or "", r["agw_address"] or "")
        a = agg.setdefault(key, [0, 0.0])
        a[0] += 1
        a[1] += _num(r["net_usd"])
    for (addr, label, agw), (cnt, usd) in agg.items():
        ws2.append([addr, label, agw, cnt, usd])
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.font = _BODY
            cell.border = _BORDER
        row[4].number_format = "$#,##0.00"
    for col, w in zip("ABCDE", [46, 16, 46, 12, 18]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # Лист 3: каталог всех обнаруженных протоколов (растёт со временем)
    ws3 = wb.create_sheet("catalog")
    ws3.append(["Протокол", "DeBank id", "Chain", "Сайт", "Впервые", "Последний раз"])
    _style_header(ws3, 6)
    for p in dao.protocol_catalog():
        ws3.append([p["name"], p["debank_id"], p["chain"], p["site_url"] or "",
                    _short_ts(p["first_seen"]), _short_ts(p["last_seen"])])
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.font = _BODY
            cell.border = _BORDER
    for col, w in zip("ABCDEF", [22, 26, 8, 40, 20, 20]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    wb.save(out_path)
    logger.ok(f"отчёт по протоколам сохранён: {out_path}")
    return out_path


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _short_ts(v) -> str:
    if not v:
        return ""
    try:
        return datetime.fromisoformat(v).strftime("%Y-%m-%d %H:%M")
    except Exception:  # noqa: BLE001
        return str(v)[:16]
