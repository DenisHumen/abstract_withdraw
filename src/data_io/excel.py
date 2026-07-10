"""Чтение wallets.xlsx и синхронизация в SQLite. Приватные ключи в БД НЕ пишутся."""
from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

from eth_account import Account
from openpyxl import Workbook, load_workbook
from web3 import Web3

from src.db.dao import Dao
from src import logger

COLUMNS = ["address", "private_key", "target_address", "proxy", "adspower_profile", "label", "enabled"]


@dataclass
class XlsxWallet:
    address: str
    private_key: str
    target_address: str | None  # None = адрес назначения ещё не задан
    proxy: str | None
    adspower_profile: str | None
    label: str | None
    enabled: bool


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def read_wallets(path: Path) -> list[XlsxWallet]:
    if not path.exists():
        raise FileNotFoundError(
            f"нет файла {path}. Создайте шаблон: python -m src.main init-data и заполните его"
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    idx = {name: header.index(name) for name in header if name}

    def cell(row: tuple, name: str) -> str | None:
        i = idx.get(name)
        if i is None or i >= len(row) or row[i] is None:
            return None
        v = str(row[i]).strip()
        return v or None

    wallets: list[XlsxWallet] = []
    for n, row in enumerate(rows[1:], start=2):
        pk = cell(row, "private_key")
        target = cell(row, "target_address")
        if not pk and not cell(row, "address"):
            continue  # пустая строка
        if not pk:
            logger.warn(f"строка {n}: нет private_key — пропуск (браузерная ветка пока не активна)")
            continue
        if not pk.startswith("0x"):
            pk = "0x" + pk
        try:
            derived = Account.from_key(pk).address
        except Exception as e:  # noqa: BLE001
            logger.error(f"строка {n}: некорректный private_key ({e}) — пропуск")
            continue

        declared = cell(row, "address")
        if declared and declared.lower() != derived.lower():
            logger.error(
                f"строка {n}: address ({declared}) не совпадает с ключом ({derived}) — пропуск"
            )
            continue
        # target_address может быть пустым: задача встанет в очередь (WAITING_TARGET)
        # и выполнится, как только адрес появится в XLSX. Если задан — валидируем.
        if target and not Web3.is_address(target):
            logger.error(f"строка {n}: target_address ({target}) невалиден — пропуск")
            continue
        if target:
            target = Web3.to_checksum_address(target)

        enabled_raw = (cell(row, "enabled") or "1").lower()
        wallets.append(
            XlsxWallet(
                address=derived,
                private_key=pk,
                target_address=target,
                proxy=cell(row, "proxy"),
                adspower_profile=cell(row, "adspower_profile"),
                label=cell(row, "label"),
                enabled=enabled_raw not in ("0", "false", "no", "нет"),
            )
        )
    return wallets


def sync_to_db(path: Path, dao: Dao) -> dict[str, str]:
    """XLSX -> SQLite. Возвращает {address: private_key} (ключи живут только в памяти)."""
    h = file_hash(path)
    wallets = read_wallets(path)
    keys: dict[str, str] = {}
    for w in wallets:
        dao.upsert_wallet(
            address=w.address,
            target_address=w.target_address,
            proxy=w.proxy,
            adspower_profile=w.adspower_profile,
            label=w.label,
            enabled=w.enabled,
        )
        keys[w.address] = w.private_key

    # XLSX — главный источник: строки, пропавшие из файла, удаляются из БД вместе с их задачами.
    # Защита: удаляем только если файл успешно распарсен и содержит хотя бы один кошелёк
    # (пустой/битый XLSX не должен обнулять базу).
    if wallets:
        removed = dao.delete_wallets_not_in([w.address for w in wallets])
        if removed:
            logger.warn(f"{removed} кошельков удалены из БД (нет в XLSX)")
    else:
        logger.warn("в XLSX нет валидных кошельков — удаление из БД пропущено (защита от обнуления)")

    if dao.get_meta("xlsx_hash") != h:
        dao.set_meta("xlsx_hash", h)
    logger.ok(f"sync: {len(wallets)} кошельков из {path.name}")
    return keys


def create_template(path: Path) -> None:
    """Шаблон wallets.xlsx для заполнения пользователем."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "wallets"
    ws.append(COLUMNS)
    # ширина колонок
    widths = [46, 70, 46, 40, 18, 16, 9]
    for col, width in zip("ABCDEFG", widths):
        ws.column_dimensions[col].width = width
    # пример-подсказка (удалить перед боевым запуском)
    ws.append(
        [
            "(опц.) 0x...adres — можно оставить пустым",
            "0x...privkey EVM-кошелька (ОБЯЗАТЕЛЬНО)",
            "0x...куда слать ETH на Base (можно пусто -> задача ждёт адрес)",
            "login:passwd@ip:port (опц.)",
            "(опц.)",
            "(опц.)",
            "1",
        ]
    )
    wb.save(path)
